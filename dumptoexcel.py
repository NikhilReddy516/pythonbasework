import click
import openpyxl
from openpyxl.utils import (get_column_letter)
from bs4 import BeautifulSoup



@click.command()
@click.argument('source_html')
@click.argument('output_excel')

def dumptoexcel(source_html, output_excel):

    """Dumps the html table data into excel. """

    arguments = {'srcName' : source_html, 'desName' :output_excel }

    #Reading from HTML file.
    soup = BeautifulSoup(open(arguments['srcName']))
    table = soup.find('table')
    table_rows = table.find_all('tr')

    
    #Opening Excel File.
    desWorkBook = openpyxl.Workbook()
    desSheet = desWorkBook.active


    #Getting data ready to write.
    all_rows = []

    table_head = table.find_all('th')
    row = [i.text for i in table_head]
    all_rows.append(row)
    
    for tr in table_rows:
        td = tr.find_all('td')
        row = [i.text for i in td]
        if(len(row) != 0):
            all_rows.append(row)

    rowLen = len(all_rows[0])
    maxColWidths = [0]*rowLen
    
    for row in all_rows:
        for i in range(0,rowLen):
            temp = len(row[i])
            if(maxColWidths[i]<temp):
               maxColWidths[i] = temp

    
    #Writing to Excel File.
    rowNo = 1
    for row in all_rows:
        colNo = 1
        row_len = len(row)
        for i in xrange(1,row_len):

            desSheet.cell(row=rowNo, column=colNo).value = row[i]
            desSheet.column_dimensions[get_column_letter(colNo)].width = maxColWidths[i]  
            colNo = colNo+1
            
        rowNo = rowNo+1

    #Saving Excel File.
        
    desWorkBook.save(arguments['desName'])
    
    


if __name__ == '__main__':
    dumptoexcel()


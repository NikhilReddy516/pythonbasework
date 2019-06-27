from openpyxl.utils import (get_column_letter)
import openpyxl
import click
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

def performCopy(arguments):
	
	srcWorkBook = openpyxl.load_workbook(arguments['srcName'])
	desWorkBook = openpyxl.Workbook()
	
	for sheetName in desWorkBook.get_sheet_names():
		desWorkBook.remove_sheet(desWorkBook.get_sheet_by_name(sheetName))
	
	
	srcSheetNames = srcWorkBook.get_sheet_names()
	
	noOfSheets = len(srcSheetNames)
	if(noOfSheets != 1):
		srcSheetNames = srcSheetNames[:noOfSheets-1]+srcSheetNames[noOfSheets:]
	
	print(srcSheetNames)
	
	for sheetName in srcSheetNames:
		
		srcSheet = srcWorkBook.get_sheet_by_name(sheetName)
		desSheet = desWorkBook.create_sheet(sheetName)
		
		rows = srcSheet.max_row
		cols = srcSheet.max_column
		
		for rowNum in xrange(1, rows+1):
			
			for colNum in xrange(1, cols+1):
				srcCell = srcSheet.cell(row=rowNum, column=colNum)
				desCell = desSheet.cell(row=rowNum, column=colNum)
				if(arguments['capitalize'] == True):
					desCell.value = srcCell.value.capitalize()
				else:
					desCell.value = srcCell.value
				
					
					#desSheet.cell(row=rowNum, column=colNum).value = srcSheet.cell(row=rowNum, column=colNum).value
					
					
				desSheet.row_dimensions[rowNum].height = srcSheet.row_dimensions[rowNum].height
				desSheet.column_dimensions[get_column_letter(colNum)].width = srcSheet.column_dimensions[get_column_letter(colNum)].width
								

				if(arguments['preserve'] == True):
					desCell.style = srcCell.style
					desCell.font = copy(srcCell.font)
					desCell.alignment = copy(srcCell.alignment)
					desCell.border = copy(srcCell.border)
					desCell.fill = copy(srcCell.fill)
					
				
				
	desWorkBook.save(arguments['desName'])
	
	
	


@click.command()
@click.option('--capitalize', default=False, is_flag = True, help='Capitalizes the string')
@click.option('--preservestyles', default=False, is_flag = True, help='Cells are copied with its styles')
@click.argument('source_excel')
@click.argument('dest_excel')

def copyexcel(capitalize, preservestyles, source_excel, dest_excel):

	"""It copies all data from one excel file to another and transforms them according to specified criteria."""
	arguments = {}
	arguments['capitalize'] = capitalize
	arguments['preserve'] = preservestyles
	arguments['srcName'] = source_excel
	arguments['desName'] = dest_excel
	
	print(preservestyles)
	print(capitalize)
	print(source_excel)
	print(dest_excel)
	
	performCopy(arguments)
	
	return
	

if(__name__ == '__main__'):
	copyexcel()
	
	
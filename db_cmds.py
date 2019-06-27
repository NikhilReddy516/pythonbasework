import pymysql
import click
import openpyxl

def createTables(cursor):

    studentQuery = """CREATE TABLE STUDENTS( 
                        NAME VARCHAR(100) NOT NULL, 
                        COLLEGE VARCHAR(30) NOT NULL, 
                        DBNAME VARCHAR(50) NOT NULL,
                        PRIMARY KEY (DBNAME)
                        );"""

    marksQuery = """CREATE TABLE MARKS (
                        NAME VARCHAR(100) NOT NULL,
                        DBNAME VARCHAR(50) NOT NULL,
                        MARKS INT NOT NULL,
                        PRIMARY KEY (NAME),
                        FOREIGN KEY (DBNAME) REFERENCES STUDENTS(DBNAME)
                        );"""

    #FOREIGN KEY (DBNAME) REFERENCES STUDENTS(DBNAME)
    cursor.execute("DROP TABLE IF EXISTS MARKS")
    cursor.execute("DROP TABLE IF EXISTS STUDENTS")

    cursor.execute(studentQuery)
    cursor.execute(marksQuery)

    foreignKey = "ALTER TABLE MARKS ADD CONSTRAINT FOREIGN KEY(DBNAME) REFERENCES STUDENTS(DBNAME)"
    cursor.execute(foreignKey)

    print("Studnets and Marks Table Created")



def getStudentInsert():

    studentWorkBook = openpyxl.load_workbook('./../assignment2-copyExcelSheets/students.xlsx')

    studentCurrSheet = studentWorkBook.get_sheet_by_name('Current')

    rows = studentCurrSheet.max_row
    cols = studentCurrSheet.max_column

    sqlString = "INSERT INTO STUDENTS VALUES ";

    for rowNum in xrange(2, rows + 1):

        sqlString += "("
        for colNum in xrange(1, cols + 1):

            if (colNum != cols - 1):
                if (colNum == cols):
                    sqlString += ("'" + (studentCurrSheet.cell(row=rowNum, column=colNum).value).lower() + "'")
                else:
                    sqlString += (("'" + studentCurrSheet.cell(row=rowNum, column=colNum).value + "'") + ",")
        sqlString += ")"
        if (rowNum != rows):
            sqlString += ","

    return sqlString


def getMarksInsert(cursor):

    marksWorkBook = openpyxl.load_workbook('./../assignment3-dumpHtmlToExcel/marks.xlsx')

    marksSheet = marksWorkBook.get_sheet_by_name('Sheet')

    rows = marksSheet.max_row
    cols = marksSheet.max_column

    sqlString = "INSERT INTO MARKS VALUES ";

    for rowNum in xrange(2, rows+1):
        name = marksSheet.cell(row=rowNum, column=1).value
        cursor.execute("SELECT DBNAME FROM STUDENTS WHERE DBNAME = '" + name.split('_')[2] + "'")
        if(cursor.fetchone()!=None):
            sqlString += "("
            for colNum in [1, 6]:
                if(colNum == 1):
                    name = marksSheet.cell(row=rowNum, column=colNum).value
                    sqlString += ("'"+name+"',")
                    sqlString += "(SELECT DBNAME FROM STUDENTS WHERE DBNAME ='"+(name.split('_'))[2]+"'),"
                    # print(rowNum, name)
                    # cursor.execute("SELECT DBNAME FROM STUDENTS WHERE DBNAME = '"+name.split('_')[2]+"'")
                    #print(cursor.fetchone())
                else:
                    sqlString += (marksSheet.cell(row=rowNum, column=colNum).value)
            sqlString += ")"

            if(rowNum != rows):
                sqlString += ","

    return sqlString


@click.group()
def start():
    """This function is used to execute operations on database."""


@start.command('createdb')
@click.argument('dbname')
def createdb(dbname):
    """It create database with given name along with students and marks tables in it."""
    try:
        db = pymysql.connect('localhost', 'root', 'root', dbname)
    except AttributeError as e:
        print ("Error while connection to database"+e)
        return

    cursor = db.cursor()

    createTables(cursor)

    print(dbname+" Database created  with student and marks tables in it.")

    db.close()



@start.command('dropdb')
@click.argument('dbName')
def dropdb(dbName):
    """It is used to drop the database with given name."""
    try:
        db = pymysql.connect('localhost', 'root', 'root', dbName)
    except:
        print("Error while connecting to database")
        return

    cursor = db.cursor()

    cursor.execute("DROP DATABASE IF EXISTS "+dbName)

    db.close()


@start.command('importdata')
@click.argument('dbname')

def importData(dbname):

    """This will copy data from excel sheets student.xlsl and marks.xlsl to database student and marks table."""

    try:
        db = pymysql.connect('localhost', 'root', 'root', dbname)
    except Error as e:
        print ("Error while connection to database"+e)
        return

    cursor = db.cursor()

    sqlString = getStudentInsert()
    cursor.execute(sqlString)
    db.commit()


    sqlString = getMarksInsert(cursor)
    cursor.execute(sqlString)
    db.commit()

    db.close()
    print("Imported Data")


@start.command('collegestat')
@click.argument('dbname')
def collegestat(dbname):

    """Used to display stats of each college with min, max and avg marks."""
    try:
        db = pymysql.connect('localhost', 'root', 'root', dbname)
    except Error as e:
        print ("Error while connection to database" + e)
        return

    cursor = db.cursor()

    cursor.execute("SELECT DISTINCT COLLEGE FROM STUDENTS")

    college_names = cursor.fetchall()

    print("College Acronym \t Number Of Students \t Max \t Min \t Avg");
    for name in college_names:
        name = name[0]
        cursor.execute("SELECT COUNT(DBNAME) FROM STUDENTS WHERE COLLEGE = '"+name+"'")
        count = cursor.fetchone()[0]
        cursor.execute("SELECT MAX(MARKS) FROM MARKS WHERE DBNAME IN (SELECT DBNAME FROM STUDENTS WHERE COLLEGE='"+name+"')")
        max = cursor.fetchone()[0]
        cursor.execute("SELECT MIN(MARKS) FROM MARKS WHERE DBNAME IN (SELECT DBNAME FROM STUDENTS WHERE COLLEGE='"+name+"')")
        min = cursor.fetchone()[0]
        cursor.execute("SELECT AVG(MARKS) FROM MARKS WHERE DBNAME IN (SELECT DBNAME FROM STUDENTS WHERE COLLEGE='" + name + "')")
        avg = cursor.fetchone()[0]

        print(name+'                            '+str(count)+'                '+str(max)+'       '+str(min)+'        '+str(avg))



if(__name__ == "__main__"):
    start()
